import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import os
import shutil
import string
import re
import openpyxl.formula.translate as translate
import traceback



# Create a Tkinter application
app = tk.Tk()
app.title("Wind Walk")
selected_column_data_assump = []
selected_column_data_assump_formulas = []
selected_sleeve_data_model = [[] for _ in range(6)]  # Create 6 empty lists
selected_column = ""
new_template_path = ""
fp_working = None
fp_template = None

# Create StringVars to hold the file paths
working_file_var = tk.StringVar()
template_file_var = tk.StringVar()
notification_var = tk.StringVar()

# Create a style for the application
style = ttk.Style()
style.configure("TButton", padding=5)
style.configure("TLabel", padding=5, font=("Helvetica", 12))

# Function to select the previous working file and display Release columns
def select_previous_file():
    global fp_working
    fp_working = filedialog.askopenfilename(title="Select Previous Working File")
    working_file_var.set(f"Previous/Working file: {fp_working}")  # Update the StringVar
    if fp_working:
        try:
            # Load the selected file
            wb = openpyxl.load_workbook(fp_working, data_only=True)
            wb_formulas = openpyxl.load_workbook(fp_working)
            
            try:
                # First, try to access the "Assump" tab
                sheet = wb["Assump"]
                sheet_formulas = wb_formulas["Assump"]
            except KeyError:
                # If "Assump" tab is not found, try accessing the "Assumptions" tab
                try:
                    sheet = wb["Assumptions"]
                    sheet_formulas = wb_formulas["Assumptions"]
                except KeyError:
                    show_error("Error", "Neither 'Assump' nor 'Assumptions' tab found.")
                    return
                
            # Find the row with "Release" and get the columns
            release_columns = []
            for col in sheet.iter_cols():
                if col[3].value == "Release":
                    release_columns.append(col[3].column_letter)

            if not release_columns:
                show_error("Error", "No Release columns found.")
                return

            # Display the available Release columns for user selection
            column_selection_window = tk.Toplevel()
            column_selection_window.title("Select Release Column")

            label = tk.Label(column_selection_window, text="Select a Release column:")
            label.pack()

            column_selection_combobox = ttk.Combobox(column_selection_window, values=release_columns)
            column_selection_combobox.pack()

            def select_column():
                global selected_column_data_assump
                global selected_column_data_assump_formulas
                global selected_column
                selected_column = column_selection_combobox.get()
                if selected_column:
                    selected_column_data_assump.clear()
                    selected_column_data_assump_formulas.clear()
                    for cell in sheet[selected_column]:
                        if hasattr(cell, 'value'):
                            selected_column_data_assump.append(cell.value)
                        else:
                            selected_column_data_assump.append(cell)
                    for cell in sheet_formulas[selected_column]:
                        if hasattr(cell, 'value'):
                            selected_column_data_assump_formulas.append(cell.value)
                        else:
                            selected_column_data_assump_formulas.append(cell)
                    selected_column_data_assump = selected_column_data_assump[3:]
                    selected_column_data_assump[0] = 'Previous'
                    selected_column_data_assump_formulas = selected_column_data_assump_formulas[3:]
                    selected_column_data_assump_formulas[0] = 'Working'

                    project_name = selected_column_data_assump[2]

                    show_info("Column Selected", f"Selected column data from Assumptions tab has been saved!")

                    if "Model" in wb.sheetnames:
                        model_sheet = wb["Model"]
                        project_index = None  # Initialize the project_index

                        # Search for the project name in the columns
                        for col_idx, column in enumerate(model_sheet.iter_cols(values_only=True)):
                            if project_name in column:
                                project_index = col_idx+1
                                break

                        if project_index is not None:
                            # Transpose the data from columns into the selected_sleeve_data_model
                            for idx, col in enumerate(model_sheet.iter_cols(min_col=project_index, max_col=project_index+5, values_only=True)):
                                selected_sleeve_data_model[idx].extend(col)
                            #print(selected_sleeve_data_model)
                            show_info("Sleeve Selected", f"Selected sleeve data from Model tab has been saved!")
                        else:
                            show_error("Error", f"Project name '{project_name}' not found in 'Model' tab.")
                    else:
                        show_error("Error", "No 'Model' tab found in the workbook.")
                else:
                    show_error("Error", "Please select a column.")

                column_selection_window.destroy()

            select_button = tk.Button(column_selection_window, text="Select", command=select_column)
            select_button.pack()
            
        except Exception as e:
            show_error("Error", str(e))

# Function to show an error message dialog
def show_error(title, message):
    messagebox.showerror(title, message)

# Function to show an information message dialog
def show_info(title, message):
    messagebox.showinfo(title, message)

    # Function to convert column index to column letter
def get_column_letter(col_idx):
    """Converts a 1-based column index to a column letter (e.g., 1 -> 'A', 2 -> 'B', 27 -> 'AA')."""
    col_letter = ''
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        col_letter = chr(65 + remainder) + col_letter
    return col_letter

# Function to select the template file
def select_template_file():
    global fp_template
    fp_template = filedialog.askopenfilename(title="Select Template File")
    template_file_var.set(f"Template file: {fp_template}")  # Update the StringVar
    if fp_template:
        try:
            wb = openpyxl.load_workbook(fp_template, data_only=True)
            sheet = wb["Assump"]

            # Check if the column with 'G' in the second row exists
            col_index = None
            for col in sheet.iter_cols(min_row=2, max_row=2):
                if col[0].value == 'G':
                    col_index = col[0].column

            if col_index is not None:
                # Check if the sixth row in that column is not blank
                if sheet.cell(row=6, column=col_index).value is not None and sheet.cell(row=6, column=col_index).value != "Paste Sleeve in Pre-Estimation Tab":
                    show_info("Pre-Estimation Check", "Pre-Estimation is filled out. Please ensure it is up to date.")
                else:
                    show_error("Pre-Estimation Check", "Pre-Estimation is missing or not up to date. Please include its most recent version in the template file.")
            else:
                show_error("Column Check", "No column with 'G' in the second row found in 'Assump' tab. Please indicate pre-estimation column with 'G' in the pre-estimation column's second row.")

        except Exception as e:
            show_error("Error", str(e))
    else:
        show_error("Error", "No template file selected.")
        
def transfer_data():
    global new_template_path
    
    if fp_template is None or not selected_column_data_assump or not selected_sleeve_data_model:
        show_error("Error", "Template file not selected or no data to transfer.")
        return

    try:
        # Create a copy of the template file with '_automated' added to the filename
        template_dir, template_filename = os.path.split(fp_template)
        template_name, template_ext = os.path.splitext(template_filename)
        new_template_filename = f"{template_name}_automated{template_ext}"
        new_template_path = os.path.join(template_dir, new_template_filename)

        # Copy the template file
        shutil.copy(fp_template, new_template_path)

        # Load the copied template file
        wb_template = openpyxl.load_workbook(new_template_path)
        sheet_template = wb_template["Assump"]

        # Find the column that contains the word 'Previous'
        target_col_index = None  # Initialize the column index
        for col_idx, col in enumerate(sheet_template.iter_cols(), start=1):
            if "Previous" in [cell.value for cell in col]:
                target_col_index = col_idx
                break

        if target_col_index is None:
            show_error("Error", "No column containing 'Previous' found in the template file's 'Assump' tab.")
            return

        # Calculate the starting column index for 'Previous' data
        start_col_index = target_col_index + 1

        # Get the starting row index
        start_row_index = 4

        # Calculate the ending column index for 'Previous' data
        end_col_index = start_col_index + 5  # Six columns after 'Previous'

        # Define the column letters corresponding to the six columns
        column_letters = ['V', 'W', 'X', 'Y', 'Z', 'AA']

        # Paste selected_column_data_assump into the 'Previous' column starting from row 4
        for idx, value in enumerate(selected_column_data_assump):
            sheet_template.cell(row=start_row_index + idx, column=target_col_index, value=value)

        # Create a dictionary to store data from rows 4 to 13 for each column
        column_data_dict = {}

        # Copy data from rows 4 to 13 for the 'Previous' column
        for idx, col_letter in enumerate(column_letters):
            col_index = start_col_index + idx
            column_data = []
            for row_idx in range(start_row_index, start_row_index + 10):
                cell = sheet_template.cell(row=row_idx, column=col_index)
                column_data.append(cell.value)
            column_data_dict[col_letter] = column_data

        # Paste selected_column_data_assump_formulas into the 6 columns after 'Previous' starting from row 4
        for idx, value in enumerate(selected_column_data_assump_formulas):
            col_index = start_col_index + idx
            if col_index <= end_col_index:  # Limit to only six columns after 'Previous'
                for row_idx in range(start_row_index + 10, len(selected_column_data_assump_formulas) + start_row_index + 10):
                    cell = sheet_template.cell(row=row_idx, column=col_index)
                    # Replace 'AC' with the current column letter
                    if selected_column in str(cell.value):
                        cell.value = cell.value.replace(selected_column, column_letters[idx])
        
        # Matches Previous and six subsequent columns with pre-estimation sleeve for specified rows
        rows_to_check = [39, 47, 116, 83, 85, 29]
        for row_idx in rows_to_check:
            for col_letter in ['U', 'V', 'W', 'X', 'Y', 'Z', 'AA']:
                col_index = sheet_template[col_letter + str(row_idx)].column
                if col_letter == 'U':
                    prev_col_index = col_index - 2  # Get the column two places before 'U' (column 'S')
                else:
                    prev_col_index = col_index - 1  # Get the column one place before the current column
                prev_col_letter = get_column_letter(prev_col_index)
                prev_cell = sheet_template[prev_col_letter + str(row_idx)]
                cell = sheet_template[col_letter + str(row_idx)]
                if cell.value != prev_cell.value:
                    cell.value = prev_cell.value
                    
        # Rename row 6 values in columns W, X, Y, Z, and AA
        row_to_rename = 6
        new_names = ['Refresh', 'Technology', 'Size', 'SS, Interconnect changes', 'Array', 'COD']

        for idx, col_letter in enumerate(column_letters):
            col_index = start_col_index + idx
            cell = sheet_template.cell(row=row_to_rename, column=col_index)
            cell.value = new_names[idx]
        '''If we want to insert new releases and test on machine that allows for external connections, allows for new format
        # Process the 'Pre-Estimation' tab to find 'Active' columns and store row 6 values in new_releases list
        sheet_pre_estimation = wb_template["Pre-Estimation"]
        new_releases = []

        # Check row 5 for 'Active' columns and save row 6 values
        for col in sheet_pre_estimation.iter_cols():
            if col[4].value == "Active":
                new_releases.append(col[5].value)

        # Process the 'Assumptions' tab to check for matching columns in row 6
        sheet_assumptions = wb_template["Assump"]

        # Create a set of existing values in row 6 of the 'Assumptions' tab
        existing_values = set(col[5].value for col in sheet_assumptions.iter_cols())

        # Iterate through new_releases and check for matches in row 6 of 'Assumptions' tab
        for release_value in new_releases:
            if release_value not in existing_values:
                # Insert a new column next to any matching column in row 6
                for idx, col in enumerate(sheet_assumptions.iter_cols(), start=1):
                    if col[5].value == release_value:
                        sheet_assumptions.insert_cols(idx + 1)
                        new_col = sheet_assumptions[idx + 1]
                        new_col[0].value = release_value
                        
                        # Insert the release_value in row 6 of the new column
                        for row in new_col:
                            row.value = release_value
                            
                        break

        # Collect values from 'Pre-Estimation' row 7 for unmatched columns and add them to estimate_requests
        estimate_requests = []

        for release_value in new_releases:
            if release_value not in existing_values:
                col_index = None
                for idx, col in enumerate(sheet_pre_estimation.iter_cols(), start=1):
                    if col[5].value == release_value:
                        col_index = idx
                        break
                if col_index is not None:
                    # Collect values from 'Pre-Estimation' row 7
                    values_row_7 = [cell.value for cell in sheet_pre_estimation[col_index] if cell.row == 7]
                    estimate_requests.append((release_value, values_row_7))

        # Insert values into the new columns in 'Assumptions' tab
        for release_value, values_row_7 in estimate_requests:
            for idx, col in enumerate(sheet_assumptions.iter_cols(), start=1):
                if col[5].value == release_value:
                    col_index = idx
                    break
            else:
                show_error("Error", f"Column for release value '{release_value}' not found.")
                continue  # Skip to the next release value if the column is not found

            for idx, value in enumerate(values_row_7):
                sheet_assumptions[col_index][idx + 1].value = value  
        '''
        # Check and update specified rows in columns U and V, delete sheet_assumptions in next line if uncommenting red
        sheet_assumptions = wb_template["Assump"] 
        for row_idx in range(30, 34):
            cell_u = sheet_assumptions['U' + str(row_idx)]
            cell_v = sheet_assumptions['V' + str(row_idx)]
            if cell_u.value != cell_v.value:
                cell_u.value = cell_v.value

        # Create a dictionary to store AA's data starting from row 8
        '''aa_data = {}
        for idx, cell in enumerate(sheet_assumptions['AA'], start=8):
            aa_data[idx] = cell.value

        # Paste AA's data into the new columns for each release value starting at row 8
        for idx, release_value in enumerate(new_releases):
            for col in sheet_assumptions.columns:  # Iterate through all columns in the worksheet
                if col[5].value == release_value:
                    col_letter = get_column_letter(col[0].column)  # Get the column letter
                    for row_idx, cell in aa_data.items():
                        col[row_idx].value = cell

                    # Replace 'AA' occurrences with the letter of the current column
                    for row_idx in range(8, len(col)):
                        if 'AA' in str(col[row_idx].value):
                            col[row_idx].value = str(col[row_idx].value).replace('AA', col_letter)'''

                              
        # Save the changes to the copied template file
        wb_template.save(new_template_path)
        show_info("Walk Completed", f"Wind walk completed successfully.\nAvailable at {new_template_path}")

    except Exception as e:
        print(f"An error occurred: {e}")
        traceback.print_exc()
        
        
# Function to update notification
def update_notification(message):
    notification_var.set(message)

# Create frames for organization
frame_select_files = ttk.Frame(app)
frame_buttons = ttk.Frame(app)
frame_notification = ttk.Frame(app)

frame_select_files.pack(pady=10)
frame_buttons.pack(pady=10)
frame_notification.pack(pady=10)


# Buttons and labels
previous_file_button = ttk.Button(frame_select_files, text="Select Previous File", command=select_previous_file)
template_file_button = ttk.Button(frame_select_files, text="Select Template File", command=select_template_file)
working_file_label = ttk.Label(frame_select_files, textvariable=working_file_var)
template_file_label = ttk.Label(frame_select_files, textvariable=template_file_var)
transfer_button = ttk.Button(frame_buttons, text="Perform Walk", command=transfer_data)
notification_label = ttk.Label(frame_notification, textvariable=notification_var)

previous_file_button.grid(row=0, column=0, padx=5, pady=5)
template_file_button.grid(row=0, column=1, padx=5, pady=5)
working_file_label.grid(row=6, column=0, padx=5, pady=5)
template_file_label.grid(row=6, column=1, padx=5, pady=5)
transfer_button.grid(row=0, column=0, padx=5, pady=5)
notification_label.grid(row=0, column=0, padx=5, pady=5)

# Start the Tkinter main loop
app.mainloop()
